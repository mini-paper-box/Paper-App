import tkinter as tk
import tkinter.font as tkFont
from tkinter import filedialog, messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap import Style
from ttkbootstrap.widgets import DateEntry
from settings import *
import openpyxl
import csv, os, shutil, logging
import win32clipboard
import win32con
from button import MainButton
from datetime import date, datetime, timedelta
from edit_form_popup import EditFormPopup
from database import MakeSimpleSql
from custom_class import CustomMeter, CustomTreeview
from dateutil import parser
from pdf_reader_singleton import get_pdf_reader
from AddressManagerPopup import AddressManagerPopup

logger = logging.getLogger(__name__)


class OrderDetails(ttk.Frame):
    def __init__(self, master, order_id, customer_id):
        super().__init__(master)
        self.columns = list(ORDER_DETAILS_COLUMN_CONFIG.keys())
        self._sort_column = None
        self._sort_reverse = False
        
        self.grid_rowconfigure(0, weight=0)  
        self.grid_rowconfigure(1, weight=0)  
        self.grid_rowconfigure(2, weight=1) 
        self.grid_rowconfigure(3, weight=0) 
        self.grid_columnconfigure(0, weight=1)
        
        self.order_id = order_id
        self.customer_id = customer_id
        self.database = MakeSimpleSql()

        # Style configuration
        style = Style('flatly')
        style.configure("TreeviewPrimary.Treeview", font=("Segoe UI", 10), rowheight=28, bootstyle='info')
        style.configure("TreeviewPrimary.Treeview.Heading", font=("Segoe UI", 11), bootstyle='info')
        style.map("TreeviewPrimary.Treeview",
                background=[("selected", "#1f6aa5")],
                foreground=[("selected", "white")])
        
        # Load data
        self.data = self.database.get_order_details(self.order_id)
        self.filtered_data = self.data.copy() if self.data else []
        self.addresses = self.database.get_order_addresses(self.order_id)
        self.suppliers = ["All"]
        self.get_suppliers()
        self.reader = get_pdf_reader()

        # Create UI components
        self.customer_panel = CustomerPanel(
            self, 
            billing_data=self.addresses.get("billing"), 
            shipping_data=self.addresses.get("shipping")
        )
        self.customer_panel.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky="ew")
        
        self.create_filters()
        self.create_widgets()
        self.create_buttons()
        self.populate_treeview()

        # Add space at the bottom
        self.blank_frame = ttk.Frame(self, height=40)
        self.blank_frame.grid(row=5, column=0, pady=5)
        self.blank_frame.propagate(False)
        
        logger.info(f"Order details loaded for order {order_id}")

    def create_widgets(self):
        """Create the main treeview widget."""
        tree_frame = ttk.Frame(self)
        tree_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=10, pady=10)
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        self.tree = CustomTreeview(tree_frame, self.columns, "extended")
        
        scrollbar_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)

        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        
        # Configure columns
        for col in self.columns:
            self.tree.heading(col, text=ORDER_DETAILS_COLUMN_CONFIG[col], 
                            command=lambda c=col: self.sort_column(c))

        # Context menu
        self.context_menu = tk.Menu(self, tearoff=0)
        self.context_menu.add_command(label="Export to Excel", command=self.export_xlsx)
        self.context_menu.add_command(label="Export to CSV", command=self.export_csv)
        self.context_menu.add_separator()
        for item in STATUS:
            self.context_menu.add_command(label=item, 
                                         command=lambda i=item: self.context_status_change(i))

        # Bind events
        self.tree.bind("<Button-3>", self.show_context_menu)
        self.tree.bind("<Double-1>", self.on_double_click)
        self.tree.bind("<Control-a>", self.select_all)

    def create_filters(self):
        """Create filter widgets with closed order options."""
        def get_last_day_of_month(date_obj):
            next_month = date_obj.replace(day=28) + timedelta(days=4)
            return next_month - timedelta(days=next_month.day)
        
        self.filter_frame = ttk.LabelFrame(self, text="Filters", padding=10)
        self.filter_frame.grid(row=1, column=0, columnspan=2, padx=10, sticky="nsew")

        # Date filters
        ttk.Label(self.filter_frame, text="Start Date").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.start_date = DateEntry(self.filter_frame, width=12)
        self.start_date.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(self.filter_frame, text="End Date").grid(row=0, column=2, padx=5, pady=5, sticky="e")
        self.end_date = DateEntry(self.filter_frame, width=12)
        self.end_date.grid(row=0, column=3, padx=5, pady=5)

        # Set default date range
        today = datetime.today()
        first_day = today.replace(day=1)
        last_day = get_last_day_of_month(today).date()
        self.end_date.entry.delete(0, "end")
        self.end_date.entry.insert(0, last_day.strftime("%Y-%m-%d")) 
        self.start_date.entry.delete(0, "end")
        self.start_date.entry.insert(0, first_day.strftime("%Y-%m-%d"))

        # Date presets
        preset_options = ["Custom", "Today", "Yesterday", "Last 7 Days", 
                         "Next 7 Days", "This Month"]
        self.preset_combo = ttk.Combobox(self.filter_frame, values=preset_options, 
                                        state="readonly", width=15)
        self.preset_combo.set("Custom")
        self.preset_combo.grid(row=0, column=4, padx=5, pady=10)

        def apply_preset(event=None):
            today = datetime.today()
            preset = self.preset_combo.get()
            if preset == "Today":
                s = e = today
            elif preset == "Yesterday":
                s = e = today - timedelta(days=1)
            elif preset == "Last 7 Days":
                s = today - timedelta(days=6)
                e = today
            elif preset == "Next 7 Days":
                s = today
                e = today + timedelta(days=6)
            elif preset == "This Month":
                s = today.replace(day=1)
                e = get_last_day_of_month(today).date()
            else:
                return
            
            self.start_date.entry.delete(0, "end")
            self.start_date.entry.insert(0, s.strftime("%Y-%m-%d")) 
            self.end_date.entry.delete(0, "end")
            self.end_date.entry.insert(0, e.strftime("%Y-%m-%d"))
            self.apply_filters()

        self.preset_combo.bind("<<ComboboxSelected>>", apply_preset)

        # Supplier filter
        ttk.Label(self.filter_frame, text="Supplier").grid(row=0, column=5, padx=5)
        self.supplier_filter = ttk.Combobox(self.filter_frame, values=self.suppliers, width=15)
        self.supplier_filter.current(0)
        self.supplier_filter.grid(row=0, column=6, padx=5)

        # Status filter
        ttk.Label(self.filter_frame, text="Status").grid(row=0, column=7, padx=5)
        self.status_filter = ttk.Combobox(self.filter_frame, 
                                         values=(["All"] + STATUS), width=12)
        self.status_filter.current(0)
        self.status_filter.grid(row=0, column=8, padx=5)

        # Filter buttons
        ttk.Button(self.filter_frame, text="Apply Filters", 
                  bootstyle="primary", command=self.apply_filters).grid(row=0, column=9, padx=10)
        ttk.Button(self.filter_frame, text="Reset", 
                  bootstyle="secondary", command=self.reset_filters).grid(row=0, column=10, padx=5)

        # Search box
        ttk.Label(self.filter_frame, text="Search:").grid(row=0, column=11, padx=(10, 0))
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(self.filter_frame, textvariable=self.search_var, width=20)
        self.search_entry.grid(row=0, column=12, padx=(0, 5))
        self.search_var.trace('w', lambda *args: self.apply_wildcard_filter())
        
        # Clear search button
        ttk.Button(self.filter_frame, text="✕", 
                  bootstyle="secondary-outline",
                  width=3,
                  command=self.clear_search).grid(row=0, column=13, padx=0)

        # Second row for closed orders options
        ttk.Label(self.filter_frame, text="Options:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        
        self.include_closed_var = tk.BooleanVar(value=False)
        self.include_closed_check = ttk.Checkbutton(
            self.filter_frame,
            text="Include Closed Items",
            variable=self.include_closed_var,
            command=self.apply_filters,
            bootstyle="primary-round-toggle"
        )
        self.include_closed_check.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="w")
        
        self.closed_only_var = tk.BooleanVar(value=False)
        self.closed_only_check = ttk.Checkbutton(
            self.filter_frame,
            text="Closed Items Only",
            variable=self.closed_only_var,
            command=self.toggle_closed_only,
            bootstyle="info-round-toggle"
        )
        self.closed_only_check.grid(row=1, column=3, columnspan=2, padx=5, pady=5, sticky="w")
        
        # Status summary
        self.status_summary_label = ttk.Label(
            self.filter_frame,
            text="",
            font=("Segoe UI", 9),
            foreground="gray"
        )
        self.status_summary_label.grid(row=1, column=5, columnspan=6, padx=5, pady=5, sticky="w")

    def clear_search(self):
        """Clear the search field."""
        self.search_var.set("")
        self.search_entry.focus()

    def toggle_closed_only(self):
        """Handle closed items only toggle."""
        if self.closed_only_var.get():
            self.include_closed_var.set(True)
            self.include_closed_check.config(state="disabled")
        else:
            self.include_closed_check.config(state="normal")
        self.apply_filters()

    def create_buttons(self):
        """Create action buttons."""
        self.btn_frame = ttk.Frame(self)
        self.btn_frame.grid(row=4, column=0, pady=5)

        for button_key, data in ORDER_DETAIL_BUTTONS.items():
            MainButton(parent=self.btn_frame,
                      col=data['col'],
                      row=data['row'],
                      text=data['text'],
                      bootstyle=data['style'],
                      func=self.main_action_press,
                      params=button_key)

    def populate_treeview(self, rows=None):
        """Populate the Treeview with data."""
        self.tree.delete(*self.tree.get_children())
        
        dataset = rows if rows is not None else self.filtered_data
        if not dataset:
            self.update_status_summary([])
            return

        for index, row in enumerate(dataset):
            dict_row = dict(row)
            values = [dict_row.get(col, "") for col in self.tree['column']]
            self.tree.insert("", "end", values=values)

        # Configure tags
        self.tree.tag_configure("even_overdue", background="#ffdddd")
        self.tree.tag_configure("odd_overdue", background="#ffcccc")
        self.tree.tag_configure("even_overdue_over_shipped", background="#ffe4e1")
        self.tree.tag_configure("odd_overdue_over_shipped", background="#ffcccc")
        self.tree.tag_configure("even_over_shipped", background="#e4f1b1")
        self.tree.tag_configure("odd_over_shipped", background="#dff395")
        
        # Auto-size columns
        self.autosize_columns()
        
        # Update summary
        self.update_status_summary(dataset)
        
        logger.debug(f"Populated treeview with {len(dataset)} rows")

    def autosize_columns(self):
        """Auto-size treeview columns based on content + header."""
        for col in self.columns:
            header_text = ORDER_DETAILS_COLUMN_CONFIG.get(col, col)
            max_width = tk.font.Font().measure(header_text)

            for item in self.tree.get_children():
                cell_text = str(self.tree.set(item, col))
                cell_width = tk.font.Font().measure(cell_text)
                if cell_width > max_width:
                    max_width = cell_width

            self.tree.column(col, width=max_width + 10)

    def apply_filters(self):
        """Apply all filters including closed items filter."""
        try:
            start_date_str = self.start_date.entry.get().strip()
            end_date_str = self.end_date.entry.get().strip()
            supplier_val = self.supplier_filter.get().strip()
            status_val = self.status_filter.get().strip()
            include_closed = self.include_closed_var.get()
            closed_only = self.closed_only_var.get()

            filtered = list(self.data)

            # Apply closed items filter first
            if closed_only:
                filtered = [r for r in filtered 
                           if dict(r).get("status", "").lower() in ["closed", "completed", "cancelled"]]
                logger.debug("Filtering for closed items only")
            elif not include_closed:
                filtered = [r for r in filtered 
                           if dict(r).get("status", "").lower() not in ["closed", "completed", "cancelled"]]
                logger.debug("Excluding closed items")

            # Apply date filter
            if start_date_str and end_date_str:
                try:
                    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                    end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                    
                    def date_in_range(row):
                        row_date_str = dict(row).get("order_date") or dict(row).get("due_date")
                        if not row_date_str:
                            return True
                        try:
                            row_date = datetime.strptime(str(row_date_str), "%Y-%m-%d").date()
                            return start_date <= row_date <= end_date
                        except:
                            return True
                    
                    filtered = [r for r in filtered if date_in_range(r)]
                except Exception as e:
                    logger.error(f"Date filter error: {e}")

            # Apply supplier filter
            if supplier_val and supplier_val != "All":
                filtered = [r for r in filtered 
                           if dict(r).get("supplier", "").lower() == supplier_val.lower()]

            # Apply status filter
            if status_val and status_val != "All" and not closed_only:
                filtered = [r for r in filtered 
                           if dict(r).get("status", "").lower() == status_val.lower()]

            self.filtered_data = filtered
            
            # Apply search if there's a search term
            search_term = self.search_var.get().strip()
            if search_term:
                self.apply_wildcard_filter()
            else:
                self.populate_treeview(filtered)
            
            logger.info(f"Filters applied: {len(filtered)} of {len(self.data)} items shown")
            
        except Exception as e:
            logger.error(f"Error applying filters: {e}", exc_info=True)
            messagebox.showerror("Filter Error", f"Error applying filters: {e}")

    def apply_wildcard_filter(self):
        """Apply search filter on top of other filters."""
        wildcard_val = self.search_var.get().strip().lower()
        
        base_data = self.filtered_data if hasattr(self, 'filtered_data') else self.data
        
        if not wildcard_val:
            self.populate_treeview(base_data)
            return
        
        def match(row):
            dict_row = dict(row)
            for col in self.columns:
                value = str(dict_row.get(col, "")).lower()
                if wildcard_val in value:
                    return True
            return False

        filtered = [r for r in base_data if match(r)]
        self.populate_treeview(filtered)
        
        logger.debug(f"Search filter applied: '{wildcard_val}' - {len(filtered)} results")

    def reset_filters(self):
        """Reset all filters to default values."""
        self.supplier_filter.current(0)
        self.status_filter.current(0)
        self.start_date.entry.delete(0, 'end')
        self.end_date.entry.delete(0, 'end')
        self.search_var.set("")
        self.include_closed_var.set(False)
        self.closed_only_var.set(False)
        self.include_closed_check.config(state="normal")
        
        self.filtered_data = list(self.data)
        self.populate_treeview(self.data)
        
        logger.info("Filters reset")

    def update_status_summary(self, rows):
        """Update the status summary label."""
        if not rows:
            self.status_summary_label.config(text="No items")
            return
        
        status_counts = {}
        for row in rows:
            status = dict(row).get("status", "Unknown")
            status_counts[status] = status_counts.get(status, 0) + 1
        
        summary_parts = []
        total = len(rows)
        summary_parts.append(f"Total: {total}")
        
        sorted_statuses = sorted(status_counts.items(), key=lambda x: x[1], reverse=True)
        for status, count in sorted_statuses[:3]:
            percentage = (count / total * 100) if total > 0 else 0
            summary_parts.append(f"{status}: {count} ({percentage:.0f}%)")
        
        summary_text = " | ".join(summary_parts)
        self.status_summary_label.config(text=summary_text)

    def get_suppliers(self):
        """Load suppliers from database."""
        suppliers = MakeSimpleSql().get_suppliers()
        if suppliers:
            for supplier in suppliers:
                self.suppliers.append(supplier[0])

    def on_double_click(self, event):
        """Handle double-click on a row."""
        selected_item = self.tree.selection()
        if selected_item:
            values = self.tree.item(selected_item[0], "values")
            logger.debug(f"Double-clicked: {values}")

    def select_all(self, event=None):
        """Select all visible rows."""
        self.tree.selection_set(self.tree.get_children())
        return "break"

    def show_context_menu(self, event):
        """Show context menu on right-click."""
        rowid = self.tree.identify_row(event.y)
        if not rowid:
            return
        
        selected = self.tree.selection()
        if rowid not in selected:
            self.tree.selection_set(rowid)
        else:
            self.tree.focus(rowid)

        self.context_menu_rowid = rowid
        self.context_menu.post(event.x_root, event.y_root)

    def context_status_change(self, command):
        """Handle status change from context menu."""
        logger.info(f"Changing status to: {command}")
        selected_rows = self.tree.selection()
        if not selected_rows:
            return

        for rowid in selected_rows:
            self.tree.set(rowid, "status", command)

        self.update_record()

    def update_record(self, rowid=None):
        """Update database records."""
        try:
            items = [rowid] if rowid else self.tree.selection()
            for item_id in items:
                values = self.tree.item(item_id)["values"]
                columns = self.tree["columns"]
                kv = dict(zip(columns, values))

                if not kv.get("order_line"):
                    continue

                MakeSimpleSql().update_po_fields(kv)
            
            self.data = self.database.get_order_details(self.order_id)
            self.apply_filters()
            
            logger.info("Records updated successfully")

        except Exception as e:
            logger.error(f"Update error: {e}", exc_info=True)
            messagebox.showerror("Error", str(e))

    def sort_column(self, col):
        """Sort treeview by column."""
        data = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        
        reverse = (self._sort_column == col and not self._sort_reverse)
        
        try:
            data.sort(key=lambda t: float(t[0]) if t[0] else 0, reverse=reverse)
        except ValueError:
            data.sort(key=lambda t: str(t[0]).lower(), reverse=reverse)

        for index, (_, k) in enumerate(data):
            self.tree.move(k, '', index)

        for c in self.tree["columns"]:
            arrow = ""
            if c == col:
                arrow = " ↓" if reverse else " ↑"
            self.tree.heading(c, text=ORDER_DETAILS_COLUMN_CONFIG.get(c, c) + arrow)

        self._sort_column = col
        self._sort_reverse = reverse

    def export_csv(self):
        """Export to CSV."""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv", 
            filetypes=[("CSV files", "*.csv")]
        )
        if not file_path:
            return

        try:
            with open(file_path, mode='w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(self.columns)
                for row_id in self.tree.get_children():
                    writer.writerow(self.tree.item(row_id)['values'])
            messagebox.showinfo("Export", "Data exported successfully.")
            logger.info(f"Exported to CSV: {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {e}")

    def export_xlsx(self):
        """Export to Excel."""
        file_path = r"C:\Users\sang.n\OneDrive - whitebird.ca\Inventory\inventory.xlsx"
        
        if not file_path or not os.path.exists(file_path):
            logger.error("Excel file not found")
            return

        try:
            wb = openpyxl.load_workbook(file_path)
        except PermissionError:
            messagebox.showerror("Error", "File is open. Please close it and try again.")
            return

        target_sheet_name = "Ready to ship"
        if target_sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=target_sheet_name)
        else:
            ws = wb[target_sheet_name]

        if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
            headers = ["Customer", "Order ID", "Docket #", "Ordered Qty", 
                      "Order Type", "Location", "Status"]
            ws.append(headers)

        for item in self.data:
            if "Dig Ink" in item.get("ink_description", ""):
                order_type = "Digital Order"
            else:
                order_type = "Brown Box"

            ws.append([
                self.addresses.get("billing", {}).get("name", ""),
                f'{item.get("order_id", "")}-{item.get("order_line", "")}',
                item.get("docket_id", ""),
                item.get("order_quantity", ""),
                order_type,
                "Unitizer",
                item.get("status", "")
            ])

        try:
            wb.save(file_path)
            messagebox.showinfo("Success", f"Exported {len(self.data)} rows")
            logger.info(f"Exported to Excel: {file_path}")
        except PermissionError:
            messagebox.showerror("Error", "Could not save — file is locked.")

    def main_action_press(self, params):
        """Handle button actions."""
        option = params
        match option:
            case "change_ship_to":
                self.open_shipto_lookup()
            case "change_bill_to":
                self.open_billto_lookup()
            case "request_qta":
                self.open_qta_form()
            case _:
                logger.warning(f"Unknown option: {option}")

    def open_shipto_lookup(self):
        AddressManagerPopup(self, self.database, customer_id=self.customer_id,
                           on_select=self.update_ship_to)

    def open_billto_lookup(self):
        AddressManagerPopup(self, self.database, customer_id=self.customer_id,
                           on_select=self.update_bill_to)
        
    def open_qta_form(self):
        from OrderEmailBuilder import OrderEmailBuilder

        ship_to = self.addresses.get("shipping")
        bill_to = self.addresses.get("billing")
        customer_name = bill_to["name"]
        order_number = self.order_id
        
        if not ship_to or not bill_to or not self.data:
            messagebox.showerror("Error", "Missing required data for email")
            return

        email_builder = OrderEmailBuilder(bill_to, ship_to, self.data)
        email_builder.verify_and_send(
            to="catherine.s@moyydesign.com",
            cc="sang.n@whitebird.ca",
            subject=f"QTA: {customer_name} - Order # {order_number}",
            attach_pdf=True
        )

    def update_ship_to(self, address_data):
        """Update shipping address."""
        self.customer_panel.shipping_form.set_data({
            "name": address_data["name_1"],
            "street": address_data["address_1"],
            "city": address_data["city"],
            "province": address_data["province"],
            "postal_code": address_data["postal_code"],
            "country": address_data["country"],
        })

        order_data = {
            "order_id": self.order_id,
            "ship_id": address_data["address_id"]
        }
        self.database.update_order(order_data)
        self.addresses = self.database.get_order_addresses(self.order_id)

    def update_bill_to(self, address_data):
        """Update billing address."""
        self.customer_panel.billing_form.set_data({
            "name": address_data["name_1"],
            "street": address_data["address_1"],
            "city": address_data["city"],
            "province": address_data["province"],
            "postal_code": address_data["postal_code"],
            "country": address_data["country"],
        })

        order_data = {
            "order_id": self.order_id,
            "bill_id": address_data["address_id"]
        }
        self.database.update_order(order_data)
        self.addresses = self.database.get_order_addresses(self.order_id)


# Customer form classes remain the same
class CustomerForm(ttk.Frame):
    def __init__(self, parent, mode="Billing", data=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.mode = mode

        ttk.Label(self, text=f"{self.mode} Information", 
                 font=("Segoe UI", 14, "bold")).grid(
            row=0, column=0, columnspan=2, pady=(0, 10)
        )

        ttk.Label(self, text="Customer Name:").grid(row=1, column=0, sticky=E, padx=5, pady=5)
        self.name_var = tk.StringVar()
        ttk.Entry(self, textvariable=self.name_var, width=40).grid(row=1, column=1, sticky=W, padx=5, pady=5)

        ttk.Label(self, text="Address", font=("Segoe UI", 12, "bold")).grid(
            row=2, column=0, columnspan=2, sticky=W, padx=5, pady=(10, 5)
        )

        self.street_var = tk.StringVar()
        self.city_var = tk.StringVar()
        self.state_var = tk.StringVar()
        self.postal_var = tk.StringVar()
        self.country_var = tk.StringVar()

        self._make_address_fields(3, self.street_var, self.city_var, 
                                 self.state_var, self.postal_var, self.country_var)
        
        if data:
            self.set_data(data)

    def _make_address_fields(self, start_row, street, city, state, postal, country):
        ttk.Label(self, text="Street:").grid(row=start_row, column=0, sticky=E, padx=5, pady=2)
        ttk.Entry(self, textvariable=street, width=40).grid(row=start_row, column=1, sticky=W, padx=5, pady=2)

        ttk.Label(self, text="City:").grid(row=start_row+1, column=0, sticky=E, padx=5, pady=2)
        ttk.Entry(self, textvariable=city, width=40).grid(row=start_row+1, column=1, sticky=W, padx=5, pady=2)

        ttk.Label(self, text="State/Province:").grid(row=start_row+2, column=0, sticky=E, padx=5, pady=2)
        ttk.Entry(self, textvariable=state, width=40).grid(row=start_row+2, column=1, sticky=W, padx=5, pady=2)

        ttk.Label(self, text="Postal Code:").grid(row=start_row+3, column=0, sticky=E, padx=5, pady=2)
        ttk.Entry(self, textvariable=postal, width=40).grid(row=start_row+3, column=1, sticky=W, padx=5, pady=2)

        ttk.Label(self, text="Country:").grid(row=start_row+4, column=0, sticky=E, padx=5, pady=2)
        ttk.Entry(self, textvariable=country, width=40).grid(row=start_row+4, column=1, sticky=W, padx=5, pady=2)

    def get_data(self):
        return {
            "name": self.name_var.get(),
            "address": {
                "street": self.street_var.get(),
                "city": self.city_var.get(),
                "state": self.state_var.get(),
                "postal": self.postal_var.get(),
                "country": self.country_var.get(),
            }
        }

    def set_data(self, data):
        self.name_var.set(data.get("name_1", ""))
        self.street_var.set(data.get("address_1", ""))
        self.city_var.set(data.get("city", ""))
        self.state_var.set(data.get("province", ""))
        self.postal_var.set(data.get("postal_code", ""))
        self.country_var.set(data.get("country", ""))


class CustomerPanel(ttk.Frame):
    def __init__(self, parent, billing_data=None, shipping_data=None, **kwargs):
        super().__init__(parent, **kwargs)
        
        # Billing Form
        self.billing_form = CustomerForm(self, mode="Billing", data=billing_data)
        self.billing_form.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        # Shipping Form
        self.shipping_form = CustomerForm(self, mode="Shipping", data=shipping_data)
        self.shipping_form.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        # Configure grid weights
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

    def get_data(self):
        """Return combined billing and shipping info."""
        return {
            "billing": self.billing_form.get_data(),
            "shipping": self.shipping_form.get_data()
        }

    def set_data(self, billing_data=None, shipping_data=None):
        """Prefill billing and shipping forms."""
        if billing_data:
            self.billing_form.set_data(billing_data)
        if shipping_data:
            self.shipping_form.set_data(shipping_data)